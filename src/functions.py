import pandas as pd
import numpy as np
import re
from itertools import combinations
from openpyxl.styles import PatternFill, Font
from openpyxl import Workbook
from scipy.stats import chi2_contingency
from statsmodels.formula.api import ols
import warnings
warnings.filterwarnings(
    "ignore", 'This pattern is interpreted as a regular expression, and has match groups.')

# %load Functions.py


def load_tool_choices(filename_tool, label_colname, keep_cols=False):
    tool_choices = pd.read_excel(
        filename_tool, sheet_name="choices", dtype="str")

    if not keep_cols:
        tool_choices = tool_choices[['list_name', 'name', label_colname]]

    # Remove rows with missing values in 'list_name' column
    tool_choices = tool_choices.dropna(subset=['list_name'])

    # Keep only distinct rows
    tool_choices = tool_choices.drop_duplicates()

    # Convert to DataFrame
    tool_choices = pd.DataFrame(tool_choices)

    return (tool_choices)


def load_tool_survey(filename_tool, label_colname, keep_cols=False):
    tool_survey = pd.read_excel(
        filename_tool, sheet_name="survey", dtype="str")

    tool_survey = tool_survey.dropna(subset=['type'])

    tool_survey['q.type'] = tool_survey['type'].apply(
        lambda x: re.split(r'\s', x)[0])
    tool_survey['list_name'] = tool_survey['type'].apply(
        lambda x: re.split(r'\s', x)[1] if re.match(r'select_', x) else None)

    # Select only relevant columns
    if not keep_cols:
        lang_code = re.split(r'::', label_colname, maxsplit=1)[1]
        lang_code = re.sub(r'\(', r'\\(', lang_code)
        lang_code = re.sub(r'\)', r'\\)', lang_code)
        cols_to_keep = tool_survey.columns[(tool_survey.columns.str.contains(f'((label)|(hint)|(constraint_message)|(required_message))::{lang_code}')) |
                                           (~tool_survey.columns.str.contains(r'((label)|(hint)|(constraint_message)|(required_message))::'))]
        tool_survey = tool_survey[cols_to_keep]

    # Find which data sheet question belongs to
    tool_survey['datasheet'] = None
    sheet_name = "main"
    for i, toolrow in tool_survey.iterrows():
        if re.search(r'begin[_ ]repeat', toolrow['type']):
            sheet_name = toolrow['name']
        elif re.search(r'end[_ ]repeat', toolrow['type']):
            sheet_name = "main"
        elif not re.search(r'((end)|(begin))[_ ]group', toolrow['type'], re.IGNORECASE):
            tool_survey.loc[i, 'datasheet'] = sheet_name

    return tool_survey


def map_names(column_name, column_values_name, summary_table, tool_survey, tool_choices,label_col, na_include=False):
    choices_shortlist = tool_choices[
        tool_choices['list_name'].values == tool_survey[tool_survey['name']
                                                        == summary_table[column_name][0]]['list_name'].values
    ][['name', label_col]]
    mapping_dict = dict(
        zip(choices_shortlist['name'], choices_shortlist[label_col]))
    if na_include is True:
        mapping_dict['No_data_available_NA'] = 'No data available (NA)'
    for value in summary_table[column_values_name]:
        if value not in mapping_dict:
            mapping_dict[value] = value
    # None breaks everything. Trying to change it
    mapping_dict['none'] = 'None '
    summary_table[column_values_name+'_orig']=summary_table[column_values_name].copy()
    summary_table[column_values_name] = summary_table[column_values_name].map(
        mapping_dict)
    return summary_table


def map_names_ls(column_name, values_list, tool_survey, tool_choices,label_col, na_include=False):
    choices_shortlist = tool_choices[
        tool_choices['list_name'].values == tool_survey[tool_survey['name']== column_name]['list_name'].values
    ][['name', label_col]]
    mapping_dict = dict(
        zip(choices_shortlist['name'], choices_shortlist[label_col]))
    if na_include is True:
        mapping_dict['No_data_available_NA'] = 'No data available (NA)'
    for value in values_list:
        if value not in mapping_dict:
            mapping_dict[value] = value
    # None breaks everything. Trying to change it
    mapping_dict['none'] = 'None '
    values_list = [mapping_dict.get(value, value) for value in values_list]
    return values_list

def weighted_mean(df, weight_column, numeric_column):
    weighted_sum = (df[numeric_column] * df[weight_column]).sum()
    total_weight = df[weight_column].sum()
    weighted_mean_result = weighted_sum / total_weight
    weighted_max_result = df[numeric_column].max()
    weighted_min_result = df[numeric_column].min()
    count = df.shape[0]
    
    sorted_df = df.sort_values(by=numeric_column)
    cum_weights = sorted_df[weight_column].cumsum()
    median_index = np.searchsorted(cum_weights, total_weight / 2.0)
    
    if cum_weights.iloc[median_index] == total_weight / 2.0 or sorted_df.shape[0] <= 2:
        weighted_median_result = sorted_df.iloc[median_index][numeric_column]
    else:
        weighted_median_result = sorted_df.iloc[median_index + 1][numeric_column]
    
    return pd.Series({'mean': weighted_mean_result,
                      'median':weighted_median_result,
                      'max': weighted_max_result,
                      'min': weighted_min_result,
                      'count': count})


def get_variable_type(data, variable_name):
    if data[variable_name].dtype == 'object':
        return 'string'
    elif data[variable_name].dtype == 'int64' or data[variable_name].dtype == 'int' or data[variable_name].dtype == 'int32':
        return 'integer'
    elif data[variable_name].dtype == 'float64' or data[variable_name].dtype == 'float' or data[variable_name].dtype == 'float32':
        return 'decimal'


def check_daf_filter(daf, data, filter_daf, tool_survey, tool_choices):
    merged_daf = filter_daf.merge(daf, on='ID', how='inner')
    # some calculate variables can be NaN
    merged_daf = merged_daf.drop(
        ['calculation', 'join', 'disaggregations'], axis=1)
    # check if rows contain NaN
    if filter_daf.isnull().values.any():
        raise ValueError("Some rows in the filter sheet contain NaN")

    # check IDs consistency
    if len(merged_daf) != len(filter_daf):
        raise ValueError("Some IDs in file are not in DAF")

    for row_id, row in merged_daf.iterrows():
        # check that filter variable are in the same sheet in the data
        if row['variable_x'] not in data[row['datasheet']].columns:
            raise ValueError(f"Filter variable {row['variable_x']} not found in {row['datasheet']}")

        value_type = type(row['value'])

        # check whether the value is an another variable
        if row["value"] in tool_survey['name'].tolist():
            # check that the variable is in the same sheet in the data
            if row['value'] not in data[row['datasheet']].columns:
                raise ValueError(f"Filter value {row['value']} not found in {row['datasheet']}")

            # check that the variable and the value have the same type
            if get_variable_type(data[row['datasheet']], row['variable_x']) != get_variable_type(data[row['datasheet']], row['value']):
                raise ValueError(f"Variable {row['variable_x']} and {row['value']} have different types")

            # check that the operation is allowed for the type
            if get_variable_type(data[row['datasheet']], row['value']) == 'string':
                if row['operation'] not in ["!=", "=="]:
                    raise ValueError(
                        f"Operation {row['operation']} not allowed for string variables")
            continue

        if value_type == str:
            # check that the variable and the value have the same type
            if get_variable_type(data[row['datasheet']], row['variable_x']) != 'string':
                raise ValueError(
                    f"Variable {row['variable_x']} has another type then filter value")
            # check that the operation is allowed for the type
            if row["operation"].strip(' ') not in ["!=", "=="]:
                raise ValueError(
                    f"Operation {row['operation']} not allowed for string variables")
        else:
            # check that the variable and the value have the same type
            if get_variable_type(data[row['datasheet']], row['variable_x']) == 'string':
                raise ValueError(
                    f"Variable {row['variable_x']} has another type then filter value")
            # check that the operation is allowed for the type
            if row["operation"].strip(' ') not in ["<", ">", "<=", ">=", "!=", "=="]:
                raise ValueError(
                    f"Operation {row['operation']} not allowed for numeric variables")


def check_daf_consistency(daf, data, sheets, resolve=False):
    # check that all variables have a datasheet
    if daf['datasheet'].isnull().values.any():
        if not resolve:
            raise ValueError('the following are missing ' +
                             ','.join(daf[daf['datasheet'].isnull().values]['variable']))
        else:
            print('the following are missing ' +
                  ','.join(daf[daf['datasheet'].isnull().values]['variable']))
            daf.dropna(subset=['datasheet'], inplace=True)

    # check that all variables in daf are in the corresponding data sheets
    for id, row in daf.iterrows():
        if row["variable"] not in data[row["datasheet"]].columns:
            if not resolve:
                raise ValueError(f"Column {row['variable']} not found in {row['datasheet']}")
            else:
                print(f"Column {row['variable']} not found in {row['datasheet']}")
                daf.drop(id, inplace=True)

        if row["disaggregations"] not in ["overall", ""] and not pd.isna(row['disaggregations']):
            row["disaggregations"] = row["disaggregations"].replace(" ", "")
            disaggregations_list = row["disaggregations"].split(",")

            for disaggregations_item in disaggregations_list:
                if disaggregations_item not in data[row["datasheet"]].columns:
                    error_message = f"Disaggregation {disaggregations_item} not found in {row['datasheet']} for variable {row['variable']}"
                    if not resolve:
                        raise ValueError(error_message)
                    else:
                        print(error_message)
                        daf.drop(id, inplace=True)
                        break

        if row["admin"] not in data[row["datasheet"]].columns:
            if not resolve:
                raise ValueError(f"admin {row['admin']} not found in {row['datasheet']} for variable {row['variable']}")
            else:
                print(f"admin {row['admin']} not found in {row['datasheet']} for variable {row['variable']}")
                daf.drop(id, inplace=True)

    # check if variables exist in more than one sheet
    sheet_dict = dict()
    for sheet in sheets:
        colnames = data[sheet].columns
        # drop from colnames the ones that are not in daf
        colnames = colnames[colnames.isin(daf['variable'])]
        sheet_dict[sheet] = set(colnames)

    # check and print all intersections
    for sheet1, sheet2 in combinations(sheet_dict.keys(), 2):
        intersection = sheet_dict[sheet1].intersection(sheet_dict[sheet2])
        if len(intersection) > 0:
            if not resolve:
                warnings.warn(f"Intersection between {sheet1} and {sheet2} : {intersection}")
            else:
                print(f"Intersection between {sheet1} and {sheet2} : {intersection}")
                # print("Resolve by removing from DAF the variables that are in both sheets")
                # daf = daf[~daf['variable'].isin(intersection)]

    for sheet in sheets:
        # check that all sheets have variables in daf
        if not sheet_dict[sheet]:
            print(f"WARNING: Sheet {sheet} has no variables in DAF")
        print(f"Sheet {sheet} has {len(sheet_dict[sheet])} variables")

    return daf


def custom_sort_key(value):
    if value in 'Total' and isinstance(value, str):
        return 'zzzzzzzzzzz'  # This super dumb but it works
    else:
        return value


def make_pivot(table, index_list, column_list, value):
    pivot_table = table.pivot_table(index=index_list,
                                    columns=column_list,
                                    values=value).reset_index()
    return pivot_table


def construct_result_table(tables_list, file_name, make_pivot_with_strata=False):
    workbook = Workbook()
    workbook.create_sheet("Table_of_content", 0)
    workbook.create_sheet("Data", 1)
    content_sheet = workbook["Table_of_content"]
    data_sheet = workbook["Data"]
    link_idx = 1

    # add columns in the content sheet
    content_sheet.append(["ID", "Link","Significance"])

    for idx, element in enumerate(tables_list):
        table, ID, label, significance = element
        if "perc" in table.columns:
            values_variable = "perc"
        elif any([x.startswith(('perc_','median_','mean_','max_','min_')) for x in table.columns]):
            values_variable = [x for x in table.columns if x.startswith(('perc_','median_','mean_','max_','min_'))]
        elif 'mean' in table.columns:
            values_variable = "mean"
        else:
            values_variable = 'category_count'
        if 'disaggregations_category_1' in table.columns:
            pivot_columns = ["disaggregations_category_1"]
        else:
            pivot_columns = []
            
        columns = [x for x in table.columns if ('disaggregations_category_' in x)]
        missed_cols = set(columns).difference(['disaggregations_category_1'])
        if len(missed_cols)>0:
            pivot_columns.extend(list(missed_cols))
            
        if values_variable == "perc" or values_variable == 'category_count':
            if make_pivot_with_strata:
                if table['admin_category'].isin(['Total']).any():
                    table_dirty = table[table['admin_category'] == 'Total']
                    table_clean = table[table['admin_category'] != 'Total']

                    pivot_table_dirty = make_pivot(
                        table_dirty, pivot_columns + ["option"], ["admin_category"], values_variable)
                    pivot_table_clean = make_pivot(
                        table_clean, pivot_columns + ["option"], ["admin_category"], values_variable)

                    pivot_table = pd.merge(pivot_table_clean, pivot_table_dirty[[
                                           'option', 'Total']], on=['option'], how='left')
                else:
                    pivot_table = make_pivot(
                        table, pivot_columns + ["option"], ["admin_category"], values_variable)
            else:
                if 'general_count' in table.columns:
                    pivot_columns.append('general_count')
                    
                pivot_table = make_pivot(
                    table, pivot_columns + ["admin_category", "full_count"], ["option"], values_variable)
                pivot_table = pivot_table.sort_values(
                    by='admin_category', key=lambda x: x.map(custom_sort_key))
        elif values_variable =='mean':
            if make_pivot_with_strata:
                # add numeric columns as a single one
                table = table.reset_index()
                ids = pivot_columns+['ID','admin_category']
                table = pd.melt(table, id_vars=ids, value_vars=['median', 'mean', 'max','min'])
                # add new columns to pivot
                values_variable = 'value'
                pivot_columns = pivot_columns +['variable']
                pivot_table = make_pivot(table, pivot_columns, ["admin_category"], values_variable)
            else:
                # if it's just a regular table - remove excessive information
                cols_to_drop = ['ID','variable','admin','disaggregations_1','total_count_perc']
                cols_to_keep = set(table.columns).difference(cols_to_drop)
                pivot_table = table[list(cols_to_keep)]
        else:
            cols_to_keep = [x for x in table.columns if 'category' in x]+['option']+\
                [x for x in table.columns if x.startswith(('perc_','median_','mean_','max_','min_'))]+[x for x in table.columns if x.endswith('_count')]
            pivot_table = table[cols_to_keep]
            
        cell_id = f"A{link_idx}"
        link_idx += len(pivot_table) + 3
        data_sheet.append([label])
        data_sheet.append(list(pivot_table.columns))
        for _, row in pivot_table.iterrows():
            if values_variable == "perc":
                row_id = data_sheet.max_row + 1
                for i, value in enumerate(row):
                    cell = data_sheet.cell(row=row_id, column=i + 1)
                    cell.value = value
                    if isinstance(value, (float, np.float64, np.float32)) and not pd.isna(value):
                        if value <= 1:
                            cell.number_format = '0.00%'
            else:
                data_sheet.append(list(row))
        data_sheet.append([])

        if isinstance(values_variable,list):
            link_value = ', '.join(values_variable)
        else:
            link_value = values_variable
            
        text_on_link = label + ' ' + link_value
        link_text = f'=HYPERLINK("#\'Data\'!{cell_id}", "{text_on_link}")'
        content_sheet.cell(row=idx + 2, column=2, value=link_text)
        content_sheet.cell(row=idx + 2, column=1, value=ID)
        content_sheet.cell(row=idx + 2, column=3, value=significance)

    for col_idx, column in enumerate(data_sheet.columns, 1):
        if col_idx == 1:
            data_sheet.column_dimensions[column[0].column_letter].width = 30
        else:
            data_sheet.column_dimensions[column[0].column_letter].width = 20

    content_sheet.column_dimensions['B'].width = 40
    for cell in content_sheet["B"][1:]:
        cell.font = Font(bold=True, color="FF0000FF")
    for cell in content_sheet["A"][1:]:
        cell.font = Font(bold=True)

    workbook.save(file_name)
    return workbook


def disaggregation_creator(daf_final, data, filter_dictionary, tool_choices, tool_survey,label_colname, check_significance, weight_column=None):

    if weight_column == None:
        for sheet in data:
            data[sheet]['weight'] = 1
        weight_column = 'weight'

    daf_final_freq = daf_final[daf_final['func'].isin(
        ['freq', 'select_one', 'select_multiple'])]
    daf_final_num = daf_final[daf_final['func'].isin(['numeric', 'mean'])]

    daf_final_freq.reset_index(inplace=True)
    daf_final_num.reset_index(inplace=True)

    df_list = []

    if len(daf_final_freq) > 0:
        for i, row in daf_final_freq.iterrows():
            # break down the disaggregations into a convenient list
            if not pd.isna(daf_final_freq.iloc[i]['disaggregations']):
                if ',' in daf_final_freq.iloc[i]['disaggregations']:
                    disaggregations = daf_final_freq.iloc[i]['disaggregations'].split(
                        ',')
                else:
                    disaggregations = [
                        daf_final_freq.iloc[i]['disaggregations']]
                disaggregations = [s.replace(" ", "") for s in disaggregations]
            else:
                disaggregations = []
            if not pd.isna(daf_final_freq.iloc[i]['calculation']):
                # break down the calculations
                if ' ' in daf_final_freq.iloc[i]['calculation']:
                    calc = daf_final_freq.iloc[i]['calculation'].split(',')
                else:
                    calc = [daf_final_freq.iloc[i]['calculation']]
                calc = [x.strip(' ') for x in calc]
            else:
                calc = 'None'

            # get the correct sheet & add filters
            if daf_final_freq.iloc[i]['ID'] in filter_dictionary.keys():
                filter_text = 'data["'+daf_final_freq.iloc[i]['datasheet'] + \
                    '"]['+filter_dictionary[daf_final_freq.iloc[i]['ID']]
                data_temp = eval(filter_text)
            else:
                data_temp = data[daf_final_freq.iloc[i]['datasheet']]

            # keep only those columns that we'll need
            selected_columns = [daf_final_freq['variable'][i]] + \
                disaggregations+[daf_final_freq['admin'][i]]+[weight_column]
            total_nrow = data_temp.shape[0]
            data_temp = data_temp[selected_columns]
            data_temp.loc[:, daf_final_freq['variable'][i]] = data_temp.loc[:, daf_final_freq['variable'][i]].apply(
                lambda x: re.sub(' +', ' ', x) if isinstance(x, str) else x)
            if 'include_na' in calc:
                data_temp.loc[:, daf_final_freq['variable'][i]] = data_temp[daf_final_freq['variable'][i]].fillna(
                    'No_data_available_NA')
                na_includer = True
            else:
                # remove NA rows
                data_temp = data_temp[data_temp[daf_final_freq['variable'][i]].notna(
                )]
                na_includer = False

            if data_temp.shape[0] > 0:
                freq_count = data_temp.shape[0]
            # keep a backup for select multiples
                data_temp_backup = data_temp.copy()

                # break down the data form SM
                if daf_final_freq.iloc[i]['q.type'] in ['select_multiple']:
                    data_temp.loc[:, daf_final_freq.iloc[i]['variable']
                                  ] = data_temp[daf_final_freq.iloc[i]['variable']].str.strip()

                    data_temp.loc[:, daf_final_freq.iloc[i]['variable']
                                  ] = data_temp[daf_final_freq.iloc[i]['variable']].str.split(' ').copy()
                    # Separate rows using explode
                    data_temp = data_temp.explode(
                        daf_final_freq.iloc[i]['variable'], ignore_index=True)

                groupby_columns = [daf_final_freq['admin'][i]] + \
                    disaggregations+[daf_final_freq['variable'][i]]
                # check significance if such was specified
                if check_significance ==True:
                    special_mapping = False
                    # check different cases of dependence              
                    if len(disaggregations)>0:
                        independent_variables = disaggregations
                        admin_variable = daf_final_freq['admin'][i]
                    elif len(disaggregations)==0 and daf_final_freq['admin'][i] not in ['Overall','overall']:
                        independent_variables = daf_final_freq['admin'][i]
                        admin_variable = 'overall'
                        special_mapping= True
                    else:
                        independent_variables = None
                        admin_variable = 'overall'
                    
                    admin_ls = data_temp[admin_variable].unique()
                    admin_frame = []
                    # quick variance analysis
                    if independent_variables is not None:
                        admin_ls = [x for x in admin_ls if x is not None]
                        
                        admin_ls = admin_ls +['general']
                        p_value_general = 1
                        all_p_values = []
                        variance_columns = [daf_final_freq['variable'][i]]+independent_variables
                    
                        for adm in admin_ls:
                            if adm != 'general':
                                data_temp_anova = data_temp[data_temp[admin_variable]==adm]
                            else:
                                data_temp_anova = data_temp

                            var_frame = data_temp_anova[variance_columns]
                            contingency_table = pd.crosstab(index = var_frame.iloc[:,0].values, columns =[var_frame[col] for col in variance_columns[1:]])
                            if not contingency_table.empty:
                                stat, p_value, dof, expected = chi2_contingency(contingency_table)
                                p_value = round(p_value,3)
                                
                                if adm == 'general':
                                    p_value_general = p_value
                                if p_value < 0.05:
                                    admin_frame = admin_frame + [adm]
                                    if adm != 'general':
                                        all_p_values = all_p_values + [p_value]
                                        
                        admin_frame = [x for x in admin_frame if x != 'general']
                        admin_frame = [x for x in admin_frame if x is not None]

                        if len(admin_frame)>0:
                            if ' Overall' in admin_frame:
                                res_frame = f'Significant relationship (pvalue={p_value_general})'
                            else:
                                if admin_variable in set(tool_survey['name']):
                                    admin_frame = map_names_ls(admin_variable,admin_frame,tool_survey, tool_choices,label_colname)                        
                                elif special_mapping==True and len(independent_variables)==1:
                                    if independent_variables[0] in set(tool_survey['name']):
                                        admin_frame = map_names_ls(independent_variables[0],admin_frame,tool_survey, tool_choices,label_colname)

                                admin_text = [f'{name} (p_value={value})' for name, value in zip(admin_frame, all_p_values)]
                                res_frame = 'Significant relationship at: '+', '.join(admin_text)
                        else:
                            res_frame = f'Insignificant relationship (pvalue={p_value_general})'
                    else:
                        res_frame = 'Not applicable'
                else:
                    res_frame = ''

                summary_stats = data_temp.groupby(groupby_columns)[
                    weight_column].agg(['sum', 'count'])
                summary_stats.rename(
                    columns={'count': 'unweighted_count'}, inplace=True)
                # get the same stats but for the full subsample (not calculating option samples)
                groupby_columns_ov = [
                    daf_final_freq['admin'][i]]+disaggregations

                summary_stats_var_om = data_temp_backup.groupby(
                    groupby_columns_ov)[weight_column].agg(['sum', 'count'])

                summary_stats.reset_index(inplace=True)
                summary_stats_var_om.reset_index(inplace=True)

                # rename them
                summary_stats.rename(
                    columns={'sum': 'weighted_count'}, inplace=True)
                summary_stats_var_om.rename(
                    columns={'sum': 'general_count'}, inplace=True)

                summary_stats_full = summary_stats.merge(
                    summary_stats_var_om, on=groupby_columns_ov, how='left')

                new_column_names = {daf_final_freq['variable'][i]: 'option',
                                    daf_final_freq['admin'][i]: 'admin_category'}

                if disaggregations != []:
                    for j, column_name in enumerate(disaggregations):
                        new_column_names[column_name] = f'disaggregations_category_{j+1}'

                summary_stats_full.rename(
                    columns=new_column_names, inplace=True)

                summary_stats_full['admin'] = daf_final_freq['admin'][i]
                summary_stats_full['variable'] = daf_final_freq['variable'][i]

                if disaggregations != []:
                    for j, column_name in enumerate(disaggregations):
                        summary_stats_full[f'disaggregations_{j+1}'] = disaggregations[j]

                # option replacer

                if tool_survey['name'].isin([daf_final_freq.loc[i, 'variable']]).any():
                    summary_stats_full = map_names(column_name='variable',
                                                   column_values_name='option',
                                                   label_col = label_colname,
                                                   summary_table=summary_stats_full,
                                                   tool_survey=tool_survey,
                                                   tool_choices=tool_choices,
                                                   na_include=na_includer)
                # disaggregations category replacer
                if disaggregations != []:
                    for j, column_name in enumerate(disaggregations):
                        if disaggregations[j] in set(tool_survey['name']):
                            summary_stats_full = map_names(column_name=f'disaggregations_{j+1}',
                                                           column_values_name=f'disaggregations_category_{j+1}',
                                                           summary_table=summary_stats_full,
                                                           label_col = label_colname,
                                                           tool_survey=tool_survey,
                                                           tool_choices=tool_choices)
                
                # admin category replacer
                if tool_survey['name'].isin([daf_final_freq.loc[i, 'admin']]).any():
                    summary_stats_full = map_names(column_name='admin',
                                                   column_values_name='admin_category',
                                                   summary_table=summary_stats_full,
                                                   label_col = label_colname,
                                                   tool_survey=tool_survey,
                                                   tool_choices=tool_choices)

                
                # add proper labels
                summary_stats_full['variable_orig'] = summary_stats_full['variable']
                summary_stats_full['variable'] = daf_final_freq['variable_label'][i]
                if disaggregations != []:
                    for j, column_name in enumerate(disaggregations):
                        disaggregations_labels = daf_final_freq['disaggregations_label'][i]
                        summary_stats_full[f'disaggregations_{j+1}_orig'] = summary_stats_full[f'disaggregations_{j+1}']
                        summary_stats_full[f'disaggregations_{j+1}'] = disaggregations_labels

                # add perc
                summary_stats_full['perc'] = round(
                    summary_stats_full['weighted_count']/summary_stats_full['general_count'], 4)

                summary_stats_full['weighted_count'] = summary_stats_full['weighted_count'].round()
                summary_stats_full['general_count'] = summary_stats_full['general_count'].round()

                if 'add_total' in calc:
                    summary_stats_total = data_temp.groupby(daf_final_freq['variable'][i])[
                        weight_column].agg(['sum','count'])  
                    summary_stats_total.rename(
                        columns={'count': 'unweighted_count'}, inplace=True)
                    summary_stats_total.reset_index(inplace=True)
                    # sometimes weights are wonky. so we're accounting for that
                    summary_stats_total['perc'] = round(
                        summary_stats_total['sum']/data_temp_backup[weight_column].sum(), 4)
                    summary_stats_total['weighted_count'] = summary_stats_total['sum'].copy().round()
                    # add count (n of non-na rows)
                    summary_stats_total['general_count'] = data_temp_backup.shape[0]
                    # drom the sum column
                    summary_stats_total.drop(columns=['sum'], inplace=True)

                    # rename columns
                    new_column_names = {
                        daf_final_freq['variable'][i]: 'option'}
                    summary_stats_total.rename(
                        columns=new_column_names, inplace=True)
                    # add new columns to match the existing format
                    summary_stats_total['admin'] = 'Total'
                    summary_stats_total['admin_category'] = 'Total'
                    summary_stats_total['variable'] = daf_final_freq['variable'][i]

                    if tool_survey['name'].isin([daf_final_freq.loc[i, 'variable']]).any():
                        summary_stats_total = map_names(column_name='variable',
                                                        column_values_name='option',
                                                        summary_table=summary_stats_total,
                                                        label_col = label_colname,
                                                        tool_survey=tool_survey,
                                                        tool_choices=tool_choices,
                                                        na_include=na_includer)
                    summary_stats_total['variable_orig'] = summary_stats_total['variable']
                    summary_stats_total['variable'] = daf_final_freq['variable_label'][i]
                    if disaggregations != []:
                        for j, column_name in enumerate(disaggregations):
                            summary_stats_total[f'disaggregations_{j+1}'] = 'Total'
                            summary_stats_total[f'disaggregations_category_{j+1}'] = 'Total'

                    summary_stats_full = pd.concat(
                        [summary_stats_full, summary_stats_total], ignore_index=True)

                summary_stats_full['full_count'] = freq_count
                if disaggregations != []:
                    label = daf_final_freq.iloc[i]['variable']+' broken down by ' + \
                        daf_final_freq.iloc[i]['disaggregations'] + \
                        ' on the admin of '+daf_final_freq.iloc[i]['admin']
                else:
                    label = daf_final_freq.iloc[i]['variable'] + \
                        ' on the admin of '+daf_final_freq.iloc[i]['admin']

                disagg_columns = [
                    col for col in summary_stats_full.columns if col.startswith('disaggregations') and not col.endswith('orig')]
                
                og_columns = [
                    col for col in summary_stats_full.columns if col.endswith('orig')]
                summary_stats_full['ID'] = daf_final_freq.iloc[i]['ID']
                summary_stats_full['total_count_perc'] = round((summary_stats_full['full_count']/total_nrow)*100,2)
                columns = ['ID', 'admin', 'admin_category', 'option',
                            'variable'] + disagg_columns + ['weighted_count','unweighted_count','perc',
                                                            'general_count', 'full_count','total_count_perc']+ og_columns
                    
                summary_stats_full = summary_stats_full[columns]
                df_list.append(
                    (summary_stats_full, daf_final_freq['ID'][i], label,res_frame))

    if len(daf_final_num) > 0:
        # Deal with numerics
        for i, row in daf_final_num.iterrows():
            if not pd.isna(daf_final_num.iloc[i]['disaggregations']):
                if ',' in daf_final_num.iloc[i]['disaggregations']:
                    disaggregations = daf_final_num.iloc[i]['disaggregations'].split(
                        ',')
                else:
                    disaggregations = [
                        daf_final_num.iloc[i]['disaggregations']]
                disaggregations = [s.replace(" ", "") for s in disaggregations]
            else:
                disaggregations = []
                
            if not pd.isna(daf_final_num.iloc[i]['calculation']):
                # break down the calculations
                if ' ' in daf_final_num.iloc[i]['calculation']:
                    calc = daf_final_num.iloc[i]['calculation'].split(',')
                else:
                    calc = [daf_final_num.iloc[i]['calculation']]
                calc = [x.strip(' ') for x in calc]
            else:
                calc = 'None'

            # get the correct sheet & add filters
            if daf_final_num.iloc[i]['ID'] in filter_dictionary.keys():
                filter_text = 'data["'+daf_final_num.iloc[i]['datasheet'] + \
                    '"]['+filter_dictionary[daf_final_num.iloc[i]['ID']]
                data_temp = eval(filter_text)
            else:
                data_temp = data[daf_final_num.iloc[i]['datasheet']]

            # keep only those columns that we'll need
            selected_columns = [daf_final_num['variable'][i]] + \
                disaggregations+[daf_final_num['admin'][i]]+[weight_column]
            data_temp = data_temp[selected_columns]
            total_nrow = data_temp.shape[0]
            # drop all NA observations
            data_temp = data_temp[data_temp[daf_final_num['variable'][i]].notna()]

            if data_temp.shape[0] > 0:
                mean_count = data_temp.shape[0]
                groupby_columns = disaggregations+[daf_final_num['admin'][i]]
                # conduct the tests around here
                if check_significance==True:
                    special_mapping = False
                    if len(disaggregations)>0:
                        independent_variables = disaggregations
                        admin_variable = daf_final_num['admin'][i]
                    elif len(disaggregations)==0 and daf_final_num['admin'][i] not in ['Overall','overall']:
                        independent_variables = [daf_final_num['admin'][i]]
                        admin_variable = 'overall'
                        special_mapping = True
                    else:
                        independent_variables = None
                        admin_variable = 'overall'
                    
                    if independent_variables is not None:
                        variance_columns = [daf_final_num['variable'][i]]+independent_variables
                        admin_ls = data_temp[daf_final_num['admin'][i]].unique()
                        admin_frame = []
                        
                        admin_ls = [x for x in admin_ls if x is not None]
                        admin_ls = admin_ls +['general']
                        
                        p_value_general = 1
                        all_p_values = []
                        
                        for adm in admin_ls:
                            if adm != 'general':
                                data_temp_anova = data_temp[data_temp[daf_final_num['admin'][i]]==adm]
                            else:
                                data_temp_anova = data_temp
                                
                            var_list = [daf_final_num['variable'][i]]+independent_variables
                            dep_list = 'C('+')+C('.join(var_list[1:len(var_list)])+')'
                            formula_mod = f'{var_list[0]} ~ {dep_list}'

                            model = ols(formula=formula_mod, data = data_temp).fit()
                            p_val = model.f_pvalue
                            p_val = round(p_val,3)
                            if adm =='general':
                                p_value_general=p_val
                            if p_val<0.05:
                                admin_frame = admin_frame + [adm]
                                if adm != 'general':
                                    all_p_values = all_p_values + [p_val]
                                
                        admin_frame = [x for x in admin_frame if x is not None]
                        admin_frame = [x for x in admin_frame if x != 'general']
                        
                        if len(admin_frame)>0:
                            
                            if ' Overall' in admin_frame:
                                res_frame_num = f'Significant relationship (pvalue={p_value_general})'
                            else:
                                if admin_variable in set(tool_survey['name']):
                                    admin_frame = map_names_ls(admin_variable,admin_frame,tool_survey, tool_choices,label_colname)
                                elif special_mapping==True and len(independent_variables)==1:
                                    if independent_variables[0] in set(tool_survey['name']):
                                        admin_frame = map_names_ls(independent_variables[0],admin_frame,tool_survey, tool_choices,label_colname)
                                        
                                admin_text = [f'{name} (p_value={value})' for name, value in zip(admin_frame, all_p_values)]
                                res_frame_num = 'Significant relationship at: '+', '.join(admin_text)
                        else:
                            res_frame_num = f'Insignificant relationship (pvalue={p_value_general})'
                    else:
                        res_frame_num = 'Not applicable'
                else:
                    res_frame_num = ''
                
                # get the general disaggregations statistics

                summary_stats = data_temp.groupby(groupby_columns).apply(
                    weighted_mean, weight_column=weight_column, numeric_column=daf_final_num['variable'][i])

                summary_stats = summary_stats.reset_index()

                new_column_names = {
                    daf_final_num['admin'][i]: 'admin_category'}

                if disaggregations != []:
                    for j, column_name in enumerate(disaggregations):
                        new_column_names[column_name] = f'disaggregations_category_{j+1}'

                summary_stats.rename(columns=new_column_names, inplace=True)

                summary_stats['admin'] = daf_final_num['admin'][i]
                summary_stats['variable'] = daf_final_num['variable'][i]

                if disaggregations != []:
                    for j, column_name in enumerate(disaggregations):
                        summary_stats[f'disaggregations_{j+1}'] = disaggregations[j]

                # disaggregations category replacer
                if disaggregations != [] and tool_survey['name'].isin(disaggregations).any():
                    for j, column_name in enumerate(disaggregations):
                        if disaggregations[j] in set(tool_survey['name']):
                            summary_stats = map_names(column_name=f'disaggregations_{j+1}',
                                                      column_values_name=f'disaggregations_category_{j+1}',
                                                      label_col = label_colname,
                                                      summary_table=summary_stats,
                                                      tool_survey=tool_survey,
                                                      tool_choices=tool_choices)

                # admin category replacer
                if tool_survey['name'].isin([daf_final_num.loc[i, 'admin']]).any():
                    summary_stats = map_names(column_name='admin',
                                              column_values_name='admin_category',
                                              label_col = label_colname,
                                              summary_table=summary_stats,
                                              tool_survey=tool_survey,
                                              tool_choices=tool_choices)

                # add proper labels
                summary_stats['variable_orig'] = summary_stats['variable']
                summary_stats['variable'] = daf_final_num['variable_label'][i]
                if disaggregations != []:
                    for j, column_name in enumerate(disaggregations):
                        disaggregations_labels = daf_final_num['disaggregations_label'][i]
                        summary_stats[f'disaggregations_{j+1}_orig'] = summary_stats[f'disaggregations_{j+1}']
                        summary_stats[f'disaggregations_{j+1}'] = disaggregations_labels

                if 'add_total' in calc:
                    summary_stats_total = weighted_mean(
                        data_temp, weight_column=weight_column, numeric_column=daf_final_num['variable'][i]).to_frame().transpose()

                    # add new columns to match the existing format
                    summary_stats_total['admin'] = 'Total'
                    summary_stats_total['admin_category'] = 'Total'
                    summary_stats_total['variable'] = daf_final_num['variable'][i]
                    
                    summary_stats_total['variable_orig'] = summary_stats_total['variable']
                    summary_stats_total['variable'] = daf_final_num['variable_label'][i]
                    if disaggregations != []:
                        for j, column_name in enumerate(disaggregations):
                            summary_stats_total[f'disaggregations_{j+1}'] = 'Total'
                            summary_stats_total[f'disaggregations_category_{j+1}'] = 'Total'

                    summary_stats = pd.concat(
                        [summary_stats, summary_stats_total], ignore_index=True)

                summary_stats['full_count'] = mean_count
                summary_stats.rename(columns = {'count' : 'weighted_count'}, inplace = True)
                if disaggregations != []:
                    label = daf_final_num.iloc[i]['variable']+' broken down by ' + \
                        daf_final_num.iloc[i]['disaggregations'] + \
                        ' on the admin of '+daf_final_num.iloc[i]['admin']
                else:
                    label = daf_final_num.iloc[i]['variable'] + \
                        ' on the admin of '+daf_final_num.iloc[i]['admin']
                summary_stats['total_count_perc'] = round((summary_stats['full_count']/total_nrow)*100,2)
                
                og_columns = [
                    col for col in summary_stats.columns if col.endswith('orig')]
                disagg_columns = [
                    col for col in summary_stats.columns if col.startswith('disaggregations') and not col.endswith('orig')]
                summary_stats['ID'] = daf_final_num.iloc[i]['ID']
                columns = ['ID', 'admin', 'admin_category', 'variable'] + \
                    disagg_columns + ['mean', 'median','min',
                                      'max', 'weighted_count', 'full_count','total_count_perc']+og_columns
                summary_stats = summary_stats[columns]

                df_list.append((summary_stats, daf_final_num['ID'][i], label,res_frame_num))
    return (df_list)

def key_creator(row):
    bit_1_gen = 'prop_'+row['q.type'] if row['q.type'] in ['select_one','select_multiple']  else 'mean'
    if 'option_orig' in row.keys():
      bit_2_option = '' if pd.isna(row['option_orig']) else f"%/%{row['option_orig']}"
    else:
        bit_2_option = 'mean'
    bit_3_admin = '@/@'+ row['admin_orig'] + '%/%' + row['admin_category_orig']
  
    cat_dem = [col for col in row.index if 'disaggregations_category' in col and col.endswith('orig')]
    cat_basic = [col for col in row.index if 'category' not in col and col.endswith('orig') and col.startswith('disaggregations')]
    
    combined_disaggs = [f"{row[basic]}%/%{row[dem]}" for basic, dem in zip(cat_basic, cat_dem) if not pd.isna(row[basic]) and not pd.isna(row[dem])]
    bit_4_disaggs = '-/-'.join(combined_disaggs)
    return bit_1_gen +'@/@' +row['variable_orig'] + bit_2_option + bit_3_admin + '-/-'+bit_4_disaggs
